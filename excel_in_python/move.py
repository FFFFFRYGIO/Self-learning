from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active

for row in range(1, 11):
    for col in range(1, 5):
        # char = chr(65 - 1 + col)
        char = get_column_letter(col)
        # print(ws[char + str(row)].value) #Letter+Number, ex: A1, B3, R120
        ws[char + str(row)].value = char + str(row)

ws.move_range("B2:D8", rows=10, cols=10)

wb.save('test3.xlsx')
