from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('test.xlsx')
ws = wb.active

for row in range(1, 11):
    for col in range(1, 5):
        # char = chr(65 - 1 + col)
        char = get_column_letter(col)
        # print(ws[char + str(row)].value)  # Letter+Number, ex: A1, B3, R120
        ws[char + str(row)].value = char + str(row)

# merging/unmarging cells:
# ws.merge_cells("A1:D1")
# ws.unmerge_cells("A1:D1")
# ws.merge_cells("B2:C3")

ws.insert_rows(7)  # new row above row 7
ws.insert_rows(7)
ws.delete_rows(10)  # Deleted row with values X8

ws.insert_cols(2)

wb.save('test.xlsx')
