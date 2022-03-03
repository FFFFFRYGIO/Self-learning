from openpyxl import Workbook, load_workbook

wb = load_workbook('Grades.xlsx')
ws = wb.active
print(ws)  # <Worksheet "Arkusz1">
print(ws['A1'])  # <Cell 'Arkusz1'.A1>
print(ws['A1'].value)  # Name

ws['A2'].value = "Test"  # Change is not saved yet
wb.save('Grades2.xlsx')  # Saving changes (now we can see changes)

print(wb.sheetnames)  # list of sheetnames

ws = wb['Arkusz2']  # we move to different sheet
print(ws)

wb.create_sheet("Test")
print(wb.sheetnames)
wb.save('Grades2.xlsx')
