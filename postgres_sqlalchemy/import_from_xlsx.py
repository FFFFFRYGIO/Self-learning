from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from create_tables import ListUser, List, Users_of_list, Task

wb = load_workbook('db_data.xlsx')

imported_lists = []
imported_users = []
imported_users_of_list = []
imported_tasks = []


#import lists
ws = wb['List']
row = 2
while True:
    name = ws[get_column_letter(1) + str(row)].value
    if name is None:
        break
    imported_lists.append(List(name=name))
    row += 1

#import users
ws = wb['User']
row = 2
while True:
    username = ws[get_column_letter(1) + str(row)].value
    if username is None:
        break
    email = ws[get_column_letter(2) + str(row)].value
    password = ws[get_column_letter(3) + str(row)].value
    imported_users.append(ListUser(username=username,email=email,password=password))
    row += 1

#import users_of_lists
ws = wb['Users_of_list']
row = 2
while True:
    user_id = ws[get_column_letter(1) + str(row)].value
    if user_id is None:
        break
    list_id = ws[get_column_letter(2) + str(row)].value
    imported_users_of_list.append(Users_of_list(user_id=user_id,list_id=list_id))
    row += 1

#import tasks
ws = wb['Task']
row = 2
while True:
    list_id = ws[get_column_letter(1) + str(row)].value
    if list_id is None:
        break
    description = ws[get_column_letter(2) + str(row)].value
    imported_tasks.append(Task(list_id=list_id,description=description))
    row += 1
